from __future__ import annotations

import argparse
import csv
import hashlib
import json
import math
import re
import shutil
import sys
from dataclasses import dataclass, asdict
from datetime import datetime
from pathlib import Path
from typing import Any
from urllib.parse import urljoin, urlparse

import openpyxl
import requests
import xlrd


ROOT = Path(__file__).resolve().parents[1]
WORK = ROOT / "work"
CACHE = WORK / "cache"
PAGES = CACHE / "pages"
FILES = CACHE / "source_files"
OUT = WORK / "data"

MOTS_BASE = "https://www.mots.go.th"
DATA_GO_PACKAGE_API = (
    "https://data.go.th/api/3/action/package_show?id=stattourism"
)
TREND_INBOUND_PACKAGE_API = (
    "https://data.go.th/api/3/action/package_show?id=trend_inbound_tourists"
)
TREND_INBOUND_CSV_URL = (
    "https://ckan.mots.go.th/dataset/445c66d8-a06a-49d9-adfc-35faca6fc785/"
    "resource/faffc63c-9507-451a-80b7-554cc0787368/download/est_2024_04_01.csv"
)
RECEIPT_VALIDATION_CATEGORIES = {
    2020: 742,
    2021: 745,
    2022: 761,
    2023: 797,
    2024: 810,
}
WORLD_BANK_API = (
    "https://api.worldbank.org/v2/country/THA/indicator/ST.INT.ARVL"
    "?format=json&per_page=80"
)

MONTH_MAP = {
    "jan": 1,
    "january": 1,
    "มค": 1,
    "มกราคม": 1,
    "feb": 2,
    "february": 2,
    "กพ": 2,
    "กุมภาพันธ์": 2,
    "mar": 3,
    "march": 3,
    "มีค": 3,
    "มีนาคม": 3,
    "apr": 4,
    "april": 4,
    "เมย": 4,
    "เมษายน": 4,
    "may": 5,
    "พค": 5,
    "พฤษภาคม": 5,
    "jun": 6,
    "june": 6,
    "มิย": 6,
    "มิถุนายน": 6,
    "jul": 7,
    "july": 7,
    "กค": 7,
    "กรกฎาคม": 7,
    "aug": 8,
    "august": 8,
    "สค": 8,
    "สิงหาคม": 8,
    "sep": 9,
    "sept": 9,
    "september": 9,
    "กย": 9,
    "กันยายน": 9,
    "oct": 10,
    "october": 10,
    "ตค": 10,
    "ตุลาคม": 10,
    "nov": 11,
    "november": 11,
    "พย": 11,
    "พฤศจิกายน": 11,
    "dec": 12,
    "december": 12,
    "ธค": 12,
    "ธันวาคม": 12,
}

MONTH_EN = {
    1: "Jan",
    2: "Feb",
    3: "Mar",
    4: "Apr",
    5: "May",
    6: "Jun",
    7: "Jul",
    8: "Aug",
    9: "Sep",
    10: "Oct",
    11: "Nov",
    12: "Dec",
}


@dataclass
class Category:
    id: int
    name: str
    parent_id: int | None
    year: int | None


@dataclass
class NewsFile:
    year: int
    category_id: int
    article_id: int
    article_nid: int
    title: str
    published: str
    link_download: str
    page_url: str
    file_url: str
    local_path: str | None = None
    sha256: str | None = None
    bytes: int | None = None
    parse_status: str | None = None
    parse_note: str | None = None
    parsed_months: int = 0
    parsed_total: int | None = None
    parsed_years: list[int] | None = None


def ensure_dirs() -> None:
    for path in [PAGES, FILES, OUT]:
        path.mkdir(parents=True, exist_ok=True)


def slugify(value: str, limit: int = 90) -> str:
    value = re.sub(r"[^\w.-]+", "_", value, flags=re.UNICODE).strip("_")
    return (value or "file")[:limit]


def cache_name(url: str) -> str:
    parsed = urlparse(url)
    suffix = Path(parsed.path).suffix or ".html"
    digest = hashlib.sha1(url.encode("utf-8")).hexdigest()[:12]
    stem = slugify(parsed.netloc + parsed.path.replace("/", "_"), 100)
    return f"{stem}_{digest}{suffix}"


def fetch_text(url: str, refresh: bool = False) -> str:
    path = PAGES / cache_name(url)
    if path.exists() and not refresh:
        return path.read_text(encoding="utf-8", errors="replace")
    response = requests.get(url, timeout=40)
    response.raise_for_status()
    text = response.text
    path.write_text(text, encoding="utf-8")
    return text


def fetch_json(url: str, refresh: bool = False) -> Any:
    path = PAGES / cache_name(url)
    if path.exists() and not refresh:
        return json.loads(path.read_text(encoding="utf-8"))
    response = requests.get(url, timeout=40)
    response.raise_for_status()
    data = response.json()
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    return data


def normalize_html(text: str) -> str:
    return (
        text.replace('\\"', '"')
        .replace("\\/", "/")
        .replace("\\u0026", "&")
        .replace("\\t", "\t")
    )


def parse_categories(html: str) -> list[Category]:
    text = normalize_html(html)
    pattern = re.compile(
        r'\{"id":(?P<id>\d+),"name":"(?P<name>[^"]*)",'
        r'"description":(?P<description>null|"[^"]*"),"slug":(?P<slug>null|"[^"]*"),'
        r'"parentId":(?P<parent>null|\d+),',
        re.S,
    )
    seen: dict[int, Category] = {}
    for match in pattern.finditer(text):
        cat_id = int(match.group("id"))
        name = match.group("name").strip()
        parent_raw = match.group("parent")
        parent_id = None if parent_raw == "null" else int(parent_raw)
        year_match = re.search(r"Tourism Statistics\s+(\d{4})", name)
        year = int(year_match.group(1)) if year_match else None
        seen[cat_id] = Category(cat_id, name, parent_id, year)
    return list(seen.values())


def parse_news_files(html: str, category_id: int, year: int, page_url: str) -> list[NewsFile]:
    text = normalize_html(html)
    pattern = re.compile(
        r'"news":\{"id":(?P<id>\d+),"nId":(?P<nid>\d+),'
        r'"title":"(?P<title>.*?)","description":.*?'
        r'"datePublished":"(?P<published>[^"]+)".*?'
        r'"linkDownload":(?P<link>null|"[^"]*")',
        re.S,
    )
    files: list[NewsFile] = []
    seen: set[tuple[int, str]] = set()
    for match in pattern.finditer(text):
        link_raw = match.group("link")
        if link_raw == "null":
            continue
        link = link_raw.strip('"')
        if not re.search(r"\.(xlsx|xls)(?:$|\?)", link, flags=re.I):
            continue
        title = match.group("title").strip()
        file_url = resolve_download_url(link)
        key = (int(match.group("id")), file_url)
        if key in seen:
            continue
        seen.add(key)
        files.append(
            NewsFile(
                year=year,
                category_id=category_id,
                article_id=int(match.group("id")),
                article_nid=int(match.group("nid")),
                title=title,
                published=match.group("published"),
                link_download=link,
                page_url=page_url,
                file_url=file_url,
            )
        )
    return files


def resolve_download_url(link: str) -> str:
    if link.startswith("http://") or link.startswith("https://"):
        return link
    link = link.lstrip("/")
    if link.startswith("download/"):
        return f"{MOTS_BASE}/{link}"
    return f"{MOTS_BASE}/images/{link}"


def try_download(url: str, destination: Path, refresh: bool = False) -> tuple[Path, bytes]:
    if destination.exists() and destination.stat().st_size > 0 and not refresh:
        return destination, destination.read_bytes()

    candidates = [url]
    if "/images/download/" in url:
        candidates.append(url.replace("/images/download/", "/download/"))
    if "/images/" in url:
        candidates.append(url.replace("/images/", "/"))

    last_error: Exception | None = None
    for candidate in dict.fromkeys(candidates):
        try:
            response = requests.get(candidate, timeout=60)
            response.raise_for_status()
            content = response.content
            if len(content) < 1000:
                raise ValueError(f"download too small: {len(content)} bytes")
            destination.write_bytes(content)
            return destination, content
        except Exception as exc:  # noqa: BLE001 - keep trying URL variants.
            last_error = exc
    raise RuntimeError(f"could not download {url}: {last_error}")


def try_download_text(url: str, destination: Path, refresh: bool = False) -> tuple[Path, str, bytes]:
    if destination.exists() and destination.stat().st_size > 0 and not refresh:
        content = destination.read_bytes()
        return destination, content.decode("utf-8-sig", errors="replace"), content
    response = requests.get(url, timeout=60)
    response.raise_for_status()
    content = response.content
    if len(content) < 1000:
        raise ValueError(f"download too small: {len(content)} bytes")
    destination.write_bytes(content)
    return destination, content.decode("utf-8-sig", errors="replace"), content


def value_to_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_label(value: Any) -> str:
    text = value_to_text(value).lower()
    text = re.sub(r"[\s\.\-_/()]+", "", text)
    return text


def parse_month(value: Any) -> int | None:
    label = normalize_label(value)
    if not label:
        return None
    if "ytd" in label or "total" in label or "รวม" in label:
        return None
    return MONTH_MAP.get(label)


def parse_year(value: Any) -> int | None:
    text = value_to_text(value)
    match = re.search(r"(25\d{2}|20\d{2})", text)
    if not match:
        return None
    year = int(match.group(1))
    if year >= 2400:
        year -= 543
    if 2012 <= year <= 2026:
        return year
    return None


def is_stop_label(value: Any) -> bool:
    label = normalize_label(value)
    return any(token in label for token in ["ytd", "total", "รวม", "change", "chang"])


def to_number(value: Any) -> int | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        if isinstance(value, float) and math.isnan(value):
            return None
        return int(round(value))
    text = str(value).strip().replace(",", "")
    if text in {"", "-", "n.a.", "NA", "na"}:
        return None
    try:
        return int(round(float(text)))
    except ValueError:
        return None


def iter_sheet_rows(path: Path) -> list[tuple[str, list[list[Any]]]]:
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        wb = openpyxl.load_workbook(path, data_only=True, read_only=False)
        sheets = []
        for name in wb.sheetnames:
            ws = wb[name]
            rows = [[cell.value for cell in row] for row in ws.iter_rows()]
            sheets.append((name, rows))
        return sheets
    if suffix == ".xls":
        book = xlrd.open_workbook(path)
        sheets = []
        for sheet in book.sheets():
            rows = []
            for r in range(sheet.nrows):
                rows.append([sheet.cell_value(r, c) for c in range(sheet.ncols)])
            sheets.append((sheet.name, rows))
        return sheets
    raise ValueError(f"unsupported workbook type: {path.suffix}")


def find_total_row(rows: list[list[Any]]) -> int | None:
    candidates: list[int] = []
    for idx, row in enumerate(rows):
        first_cells = row[:4]
        labels = [normalize_label(value) for value in first_cells]
        if any(label in {"grandtotal", "grandtotals"} for label in labels):
            return idx
        if any("รวมทั้งสิ้น" in label or label == "total" for label in labels):
            candidates.append(idx)
    return candidates[-1] if candidates else None


def choose_header_row(rows: list[list[Any]], total_row: int) -> int | None:
    best: tuple[int, int, int] | None = None
    for idx in range(0, total_row):
        months = [parse_month(value) for value in rows[idx]]
        unique_months = {month for month in months if month is not None}
        score = len(unique_months)
        if score < 2:
            continue
        jan_col = next((col for col, month in enumerate(months) if month == 1), 999)
        key = (score, -jan_col, idx)
        if best is None or key > best:
            best = key
    return best[2] if best else None


def extract_monthly_from_workbook(path: Path) -> tuple[dict[int, int], str]:
    best_months: dict[int, int] = {}
    best_note = ""
    sheet_months: dict[int, int] = {}
    sheet_notes: list[str] = []
    for sheet_name, rows in iter_sheet_rows(path):
        total_row = find_total_row(rows)
        if total_row is None:
            continue
        sheet_month = parse_month(sheet_name)
        if sheet_month is not None:
            total_values = [to_number(value) for value in rows[total_row][1:]]
            total_number = next((value for value in total_values if value is not None), None)
            if total_number is not None:
                sheet_months[sheet_month] = total_number
                sheet_notes.append(f"{sheet_name}:total_row={total_row + 1}")
        header_row = choose_header_row(rows, total_row)
        if header_row is None:
            continue
        header = rows[header_row]
        total = rows[total_row]
        months: dict[int, int] = {}
        started = False
        for col, header_value in enumerate(header):
            month = parse_month(header_value)
            if month is None:
                if started and is_stop_label(header_value):
                    break
                continue
            if started and month in months:
                break
            if not started and month != 1:
                continue
            started = True
            if col >= len(total):
                continue
            number = to_number(total[col])
            if number is not None:
                months[month] = number
        note = f"sheet={sheet_name}; header_row={header_row + 1}; total_row={total_row + 1}"
        if len(months) > len(best_months):
            best_months = months
            best_note = note
    if len(sheet_months) > len(best_months):
        best_months = dict(sorted(sheet_months.items()))
        best_note = "monthly_sheets; " + "; ".join(sheet_notes[:4])
    if not best_months:
        raise ValueError("could not find Grand Total monthly row")
    return dict(sorted(best_months.items())), best_note


def extract_multi_year_monthly_from_workbook(path: Path) -> tuple[dict[int, dict[int, int]], str] | None:
    for sheet_name, rows in iter_sheet_rows(path):
        for header_idx, row in enumerate(rows):
            year_cols: dict[int, int] = {}
            for col, value in enumerate(row):
                year = parse_year(value)
                if year is not None:
                    year_cols[col] = year
            if len(year_cols) < 2:
                continue
            parsed: dict[int, dict[int, int]] = {year: {} for year in year_cols.values()}
            for body_idx in range(header_idx + 1, min(len(rows), header_idx + 25)):
                body = rows[body_idx]
                month = parse_month(body[0] if body else None)
                if month is None:
                    if parsed and any(parsed.values()) and is_stop_label(body[0] if body else None):
                        break
                    continue
                for col, year in year_cols.items():
                    if col >= len(body):
                        continue
                    number = to_number(body[col])
                    if number is not None:
                        parsed.setdefault(year, {})[month] = number
            parsed = {year: dict(sorted(months.items())) for year, months in parsed.items() if len(months) >= 6}
            if parsed:
                note = f"multi_year_table; sheet={sheet_name}; header_row={header_idx + 1}"
                return parsed, note
    return None


def extract_year_blocks_from_grand_total_table(path: Path) -> tuple[dict[int, dict[int, int]], str] | None:
    for sheet_name, rows in iter_sheet_rows(path):
        total_row = find_total_row(rows)
        if total_row is None:
            continue
        header_row = choose_header_row(rows, total_row)
        if header_row is None or header_row == 0:
            continue
        month_header = rows[header_row]
        year_header = rows[header_row - 1]
        total = rows[total_row]
        parsed: dict[int, dict[int, int]] = {}
        year_starts = [
            (col, parse_year(value))
            for col, value in enumerate(year_header)
            if parse_year(value) is not None
        ]
        for i, (start_col, year) in enumerate(year_starts):
            assert year is not None
            end_col = year_starts[i + 1][0] if i + 1 < len(year_starts) else len(month_header)
            months: dict[int, int] = {}
            for col in range(start_col, end_col):
                month = parse_month(month_header[col] if col < len(month_header) else None)
                if month is None:
                    if months and is_stop_label(month_header[col] if col < len(month_header) else None):
                        break
                    continue
                if month in months:
                    break
                if col >= len(total):
                    continue
                number = to_number(total[col])
                if number is not None:
                    months[month] = number
            if len(months) >= 3:
                parsed[year] = dict(sorted(months.items()))
        if parsed:
            note = f"year_blocks; sheet={sheet_name}; year_row={header_row}; header_row={header_row + 1}; total_row={total_row + 1}"
            return parsed, note
    return None


def extract_workbook_monthlies(path: Path, default_year: int) -> tuple[dict[int, dict[int, int]], str]:
    year_blocks = extract_year_blocks_from_grand_total_table(path)
    if year_blocks is not None:
        return year_blocks
    multi = extract_multi_year_monthly_from_workbook(path)
    if multi is not None:
        return multi
    months, note = extract_monthly_from_workbook(path)
    return {default_year: months}, note


def extract_annual_total_arrivals(path: Path) -> tuple[int, str]:
    candidates: list[tuple[int, int, str]] = []
    for sheet_name, rows in iter_sheet_rows(path):
        total_row = find_total_row(rows)
        if total_row is None:
            continue
        row = rows[total_row]
        for col in range(1, min(len(row), 4)):
            number = to_number(row[col])
            if number is not None:
                sheet_key = re.sub(r"[^A-Z0-9]+", " ", str(sheet_name).upper()).strip()
                score = 0
                if "Q1 Q4" in sheet_key or "Q1Q4" in sheet_key:
                    score += 100
                if "SUMMARY" in sheet_key or "สรุป" in str(sheet_name):
                    score += 80
                if re.fullmatch(r"Q[1-4]", sheet_key):
                    score -= 50
                note = f"annual_total; sheet={sheet_name}; total_row={total_row + 1}; col={col + 1}"
                candidates.append((score, number, note))
                break
    if candidates:
        score, number, note = max(candidates, key=lambda item: (item[0], item[1]))
        return number, f"{note}; sheet_score={score}"
    raise ValueError("could not find annual Grand Total arrivals")


def year_from_title(name: str) -> int | None:
    match = re.search(r"Tourism Statistics\s+(\d{4})", name)
    return int(match.group(1)) if match else None


def parse_datetime(value: str) -> datetime:
    try:
        return datetime.fromisoformat(value.replace("Z", "+00:00"))
    except ValueError:
        return datetime.min


def crawl_mots(refresh: bool = False) -> tuple[list[NewsFile], dict[str, Any]]:
    category_411 = fetch_text(f"{MOTS_BASE}/news/category/411", refresh=refresh)
    all_categories = parse_categories(category_411)
    year_categories = [
        cat
        for cat in all_categories
        if cat.parent_id == 411 and cat.year is not None and 2012 <= cat.year <= 2026
    ]
    year_categories = sorted({cat.year: cat for cat in year_categories}.values(), key=lambda c: c.year)

    source_files: list[NewsFile] = []
    category_map: dict[int, Category] = {cat.id: cat for cat in all_categories}
    for cat in year_categories:
        year_page_url = f"{MOTS_BASE}/news/category/{cat.id}"
        year_html = fetch_text(year_page_url, refresh=refresh)
        cats = parse_categories(year_html)
        for subcat in cats:
            category_map[subcat.id] = subcat
        intl_categories = [
            subcat
            for subcat in cats
            if subcat.parent_id == cat.id
            and is_arrivals_category(subcat.name)
        ]
        scan_categories = intl_categories or [cat]
        for scan_cat in scan_categories:
            page_url = f"{MOTS_BASE}/news/category/{scan_cat.id}"
            html = fetch_text(page_url, refresh=refresh)
            news_files = parse_news_files(html, scan_cat.id, cat.year or 0, page_url)
            source_files.extend(news_files)

    # MOTS keeps a revised 2016-2019 monthly foreign-tourist workbook under
    # the 2019 category. It is a source of counts, not receipts, and repairs
    # old-year pages where arrivals were later consolidated.
    consolidated_url = f"{MOTS_BASE}/news/category/585"
    consolidated_html = fetch_text(consolidated_url, refresh=refresh)
    source_files.extend(parse_news_files(consolidated_html, 585, 2019, consolidated_url))

    data_go = fetch_json(DATA_GO_PACKAGE_API, refresh=refresh)
    trend_inbound = fetch_json(TREND_INBOUND_PACKAGE_API, refresh=refresh)
    metadata = {
        "mots_category_url": f"{MOTS_BASE}/news/category/411",
        "data_go_package_api": DATA_GO_PACKAGE_API,
        "data_go_result": data_go.get("result", {}),
        "trend_inbound_package_api": TREND_INBOUND_PACKAGE_API,
        "trend_inbound_result": trend_inbound.get("result", {}),
        "trend_inbound_csv_url": TREND_INBOUND_CSV_URL,
        "world_bank_api": WORLD_BANK_API,
        "crawl_at": datetime.now().isoformat(timespec="seconds"),
    }
    return source_files, metadata


def is_arrivals_category(name: str) -> bool:
    normalized = name.lower()
    excluded = [
        "receipt",
        "expenditure",
        "airport",
        "traveler statistics",
        "รายได้",
        "ค่าใช้จ่าย",
        "รายจ่าย",
        "สนามบิน",
        "ผู้เดินทางเข้า - ออก",
    ]
    if any(term in normalized for term in excluded):
        return False
    included = [
        "international tourist arrivals to thailand",
        "สถิตินักท่องเที่ยวชาวต่างชาติ",
        "จำนวนนักท่องเที่ยวต่างชาติ รายเดือน",
    ]
    return any(term in normalized for term in included)


def download_and_parse(files: list[NewsFile], refresh: bool = False) -> tuple[list[dict[str, Any]], list[NewsFile]]:
    monthly_rows: list[dict[str, Any]] = []
    parsed_files: list[NewsFile] = []

    for source in files:
        suffix = Path(urlparse(source.file_url).path).suffix.lower() or ".xlsx"
        filename = f"{source.year}_{source.category_id}_{source.article_id}_{slugify(source.title, 60)}{suffix}"
        local = FILES / filename
        try:
            local, content = try_download(source.file_url, local, refresh=refresh)
            source.local_path = str(local.relative_to(ROOT))
            source.sha256 = hashlib.sha256(content).hexdigest()
            source.bytes = len(content)
            parsed_by_year, note = extract_workbook_monthlies(local, source.year)
            source.parse_status = "ok"
            source.parse_note = note
            source.parsed_years = sorted(parsed_by_year)
            source.parsed_months = max(len(months) for months in parsed_by_year.values())
            source.parsed_total = sum(sum(months.values()) for months in parsed_by_year.values())
            for parsed_year, months in parsed_by_year.items():
                for month, arrivals in months.items():
                    monthly_rows.append(
                        {
                            "year": parsed_year,
                            "month": month,
                            "date": f"{parsed_year}-{month:02d}-01",
                            "arrivals": arrivals,
                            "source_article_id": source.article_id,
                            "source_year_month_count": len(months),
                            "source_title": source.title,
                            "source_published": source.published,
                            "source_page_url": source.page_url,
                            "source_file_url": source.file_url,
                            "source_local_file": source.local_path,
                            "source_sha256": source.sha256,
                            "parse_note": note,
                        }
                    )
        except Exception as exc:  # noqa: BLE001 - record parse failures for audit.
            source.parse_status = "error"
            source.parse_note = str(exc)
        parsed_files.append(source)

    monthly_rows = sorted(monthly_rows, key=lambda row: (row["year"], row["month"], row["source_article_id"]))
    return monthly_rows, parsed_files


def select_best_monthly_rows(monthly_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    # Keep the best source per year and month. The best file has the largest
    # number of months for its year, then newest publication date.
    source_scores: dict[tuple[int, int], tuple[int, datetime, int]] = {}
    for row in monthly_rows:
        key = (int(row["year"]), int(row["source_article_id"]))
        source_scores.setdefault(
            key,
            (
                int(row["source_year_month_count"]),
                parse_datetime(str(row["source_published"])),
                0,
            ),
        )
        score = source_scores[key]
        source_scores[key] = (score[0], score[1], score[2] + int(row["arrivals"]))

    best_source_by_year: dict[int, int] = {}
    for (year, article_id), score in source_scores.items():
        existing_article = best_source_by_year.get(year)
        if existing_article is None:
            best_source_by_year[year] = article_id
            continue
        if score > source_scores[(year, existing_article)]:
            best_source_by_year[year] = article_id

    final_rows = [
        row
        for row in monthly_rows
        if int(row["source_article_id"]) == best_source_by_year.get(int(row["year"]))
    ]
    final_rows = sorted(final_rows, key=lambda row: (row["year"], row["month"]))
    return final_rows


def load_trend_inbound_monthlies(refresh: bool = False) -> tuple[list[dict[str, Any]], NewsFile]:
    local = FILES / "data_go_trend_inbound_tourists_2015_2023.csv"
    local, text, content = try_download_text(TREND_INBOUND_CSV_URL, local, refresh=refresh)
    rows_by_month: dict[tuple[int, int], int] = {}
    reader = csv.DictReader(text.splitlines())
    number_col = next((name for name in reader.fieldnames or [] if "Number" in name), None)
    if number_col is None:
        raise ValueError("trend inbound CSV has no Number column")
    for row in reader:
        date_text = value_to_text(row.get("date"))
        parts = date_text.split("/")
        if len(parts) != 3:
            continue
        # Source dates are day/month/year, for example 1/12/2023.
        _day, month_raw, year_raw = parts
        year = int(year_raw)
        month = int(month_raw)
        raw = value_to_text(row.get(number_col))
        clean = re.sub(r"[^0-9-]", "", raw)
        number = 0 if clean in {"", "-"} else int(clean)
        rows_by_month[(year, month)] = rows_by_month.get((year, month), 0) + number

    source = NewsFile(
        year=2023,
        category_id=0,
        article_id=900001,
        article_nid=0,
        title="data.go.th / MOTS: trend_inbound_tourists monthly country-level CSV",
        published="2024-04-02T00:00:00+00:00",
        link_download=TREND_INBOUND_CSV_URL,
        page_url="https://data.go.th/dataset/trend_inbound_tourists",
        file_url=TREND_INBOUND_CSV_URL,
        local_path=str(local.relative_to(ROOT)),
        sha256=hashlib.sha256(content).hexdigest(),
        bytes=len(content),
        parse_status="ok",
        parse_note="country_month_csv; summed all countries by month; source dates are day/month/year",
        parsed_months=12,
        parsed_total=sum(rows_by_month.values()),
        parsed_years=sorted({year for year, _month in rows_by_month}),
    )
    monthly_rows = []
    counts_by_year: dict[int, int] = {}
    for year, month in rows_by_month:
        counts_by_year[year] = counts_by_year.get(year, 0) + 1
    for (year, month), arrivals in sorted(rows_by_month.items()):
        monthly_rows.append(
            {
                "year": year,
                "month": month,
                "date": f"{year}-{month:02d}-01",
                "arrivals": arrivals,
                "source_article_id": source.article_id,
                "source_year_month_count": counts_by_year[year],
                "source_title": source.title,
                "source_published": source.published,
                "source_page_url": source.page_url,
                "source_file_url": source.file_url,
                "source_local_file": source.local_path,
                "source_sha256": source.sha256,
                "parse_note": source.parse_note,
            }
        )
    return monthly_rows, source


def load_annual_extras(refresh: bool = False) -> tuple[list[dict[str, Any]], list[NewsFile]]:
    page_url = f"{MOTS_BASE}/news/category/549"
    html = fetch_text(page_url, refresh=refresh)
    sources = parse_news_files(html, 549, 2012, page_url)
    annual_rows: list[dict[str, Any]] = []
    parsed_sources: list[NewsFile] = []
    for source in sources:
        if "EXPENDITURE ITEM" in source.title.upper() or "ค่าใช้จ่ายเฉลี่ย" in source.title:
            continue
        suffix = Path(urlparse(source.file_url).path).suffix.lower() or ".xlsx"
        filename = f"{source.year}_{source.category_id}_{source.article_id}_{slugify(source.title, 60)}{suffix}"
        local = FILES / filename
        try:
            local, content = try_download(source.file_url, local, refresh=refresh)
            arrivals, note = extract_annual_total_arrivals(local)
            source.local_path = str(local.relative_to(ROOT))
            source.sha256 = hashlib.sha256(content).hexdigest()
            source.bytes = len(content)
            source.parse_status = "ok"
            source.parse_note = note
            source.parsed_years = [2012]
            source.parsed_months = 0
            source.parsed_total = arrivals
            annual_rows.append(
                {
                    "year": 2012,
                    "date": "2012-01-01",
                    "arrivals": arrivals,
                    "months": 0,
                    "is_full_year": True,
                    "annual_only": True,
                    "source_title": source.title,
                    "source_published": source.published,
                    "source_page_url": source.page_url,
                    "source_file_url": source.file_url,
                    "source_sha256": source.sha256,
                    "parse_note": note,
                }
            )
        except Exception as exc:  # noqa: BLE001 - keep audit trail.
            source.parse_status = "error"
            source.parse_note = str(exc)
        parsed_sources.append(source)
    return annual_rows, parsed_sources


def is_receipt_summary_source(source: NewsFile) -> bool:
    searchable = f"{source.title} {source.file_url}".upper().replace("_", " ")
    if "EXPENDITURE ITEM" in searchable or "BY EXPE" in searchable:
        return False
    if "ค่าใช้จ่ายเฉลี่ย" in source.title:
        return False
    return "TOURISM RECEIPTS FROM INTERNATIONAL TOURIST ARRIVALS" in searchable


def load_receipt_annual_validations(refresh: bool = False) -> tuple[dict[int, dict[str, Any]], list[NewsFile]]:
    validations: dict[int, dict[str, Any]] = {}
    parsed_sources: list[NewsFile] = []
    for year, category_id in RECEIPT_VALIDATION_CATEGORIES.items():
        page_url = f"{MOTS_BASE}/news/category/{category_id}"
        html = fetch_text(page_url, refresh=refresh)
        sources = [source for source in parse_news_files(html, category_id, year, page_url) if is_receipt_summary_source(source)]
        candidate_validations: list[dict[str, Any]] = []
        sources = sorted(sources, key=lambda source: parse_datetime(source.published), reverse=True)
        for source in sources:
            suffix = Path(urlparse(source.file_url).path).suffix.lower() or ".xlsx"
            filename = f"{source.year}_{source.category_id}_{source.article_id}_{slugify(source.title, 60)}{suffix}"
            local = FILES / filename
            try:
                local, content = try_download(source.file_url, local, refresh=refresh)
                arrivals, note = extract_annual_total_arrivals(local)
                source.local_path = str(local.relative_to(ROOT))
                source.sha256 = hashlib.sha256(content).hexdigest()
                source.bytes = len(content)
                source.parse_status = "ok"
                source.parse_note = note
                source.parsed_years = [year]
                source.parsed_months = 0
                source.parsed_total = arrivals
                candidate_validations.append({
                    "year": year,
                    "arrivals": arrivals,
                    "source_title": source.title,
                    "source_published": source.published,
                    "source_page_url": source.page_url,
                    "source_file_url": source.file_url,
                    "source_sha256": source.sha256,
                    "parse_note": note,
                })
            except Exception as exc:  # noqa: BLE001 - keep audit trail.
                source.parse_status = "error"
                source.parse_note = str(exc)
            parsed_sources.append(source)
        if candidate_validations:
            # Some MOTS categories include corrected or partial workbooks with the same broad title.
            # The annual summary total should be the full-country total, so choose the largest parsed
            # arrivals count among valid non-expenditure summary workbooks for the year.
            validations[year] = max(candidate_validations, key=lambda item: item["arrivals"])
    return validations, parsed_sources


def load_world_bank(refresh: bool = False) -> dict[int, int]:
    data = fetch_json(WORLD_BANK_API, refresh=refresh)
    values: dict[int, int] = {}
    if isinstance(data, list) and len(data) > 1:
        for item in data[1]:
            value = item.get("value")
            if value is not None:
                values[int(item["date"])] = int(value)
    return values


def build_derived(
    monthly_rows: list[dict[str, Any]],
    wb_values: dict[int, int],
    annual_extras: list[dict[str, Any]] | None = None,
    receipt_validations: dict[int, dict[str, Any]] | None = None,
) -> dict[str, Any]:
    by_year: dict[int, list[dict[str, Any]]] = {}
    for row in monthly_rows:
        by_year.setdefault(int(row["year"]), []).append(row)

    monthly = []
    previous_arrivals: int | None = None
    by_year_month = {(int(r["year"]), int(r["month"])): int(r["arrivals"]) for r in monthly_rows}
    for row in sorted(monthly_rows, key=lambda r: (r["year"], r["month"])):
        year = int(row["year"])
        month = int(row["month"])
        arrivals = int(row["arrivals"])
        yoy_base = by_year_month.get((year - 1, month))
        mom = pct_change(arrivals, previous_arrivals)
        yoy = pct_change(arrivals, yoy_base)
        monthly.append({**row, "mom_pct": mom, "yoy_pct": yoy, "month_name": MONTH_EN[month]})
        previous_arrivals = arrivals

    quarterly = []
    for year, rows in by_year.items():
        months = {int(row["month"]): int(row["arrivals"]) for row in rows}
        for quarter in range(1, 5):
            q_months = [(quarter - 1) * 3 + m for m in range(1, 4)]
            if all(month in months for month in q_months):
                total = sum(months[month] for month in q_months)
                quarterly.append(
                    {
                        "year": year,
                        "quarter": quarter,
                        "period": f"Q{quarter}",
                        "date": f"{year}-{q_months[0]:02d}-01",
                        "arrivals": total,
                    }
                )
    q_lookup = {(row["year"], row["quarter"]): row["arrivals"] for row in quarterly}
    quarterly = sorted(quarterly, key=lambda r: (r["year"], r["quarter"]))
    previous_q: int | None = None
    for row in quarterly:
        row["qoq_pct"] = pct_change(row["arrivals"], previous_q)
        row["yoy_pct"] = pct_change(row["arrivals"], q_lookup.get((row["year"] - 1, row["quarter"])))
        previous_q = row["arrivals"]

    annual = []
    for year, rows in sorted(by_year.items()):
        total = sum(int(row["arrivals"]) for row in rows)
        month_count = len(rows)
        annual.append(
            {
                "year": year,
                "date": f"{year}-01-01",
                "arrivals": total,
                "months": month_count,
                "is_full_year": month_count == 12,
                "annual_only": False,
                "world_bank_arrivals": wb_values.get(year),
            }
        )
    existing_annual_years = {row["year"] for row in annual}
    for extra in annual_extras or []:
        if extra["year"] in existing_annual_years:
            continue
        annual.append({**extra, "world_bank_arrivals": wb_values.get(extra["year"])})
    annual = sorted(annual, key=lambda row: row["year"])
    annual_lookup = {row["year"]: row["arrivals"] for row in annual}
    for row in annual:
        if row["months"] not in (0, 12):
            comparison_months = range(1, int(row["months"]) + 1)
            base_values = [by_year_month.get((row["year"] - 1, month)) for month in comparison_months]
            yoy_base = sum(base_values) if all(value is not None for value in base_values) else None
            row["yoy_basis"] = f"YTD same {row['months']} months"
        else:
            yoy_base = annual_lookup.get(row["year"] - 1)
            row["yoy_basis"] = "full_year"
        row["yoy_pct"] = pct_change(row["arrivals"], yoy_base)
        wb_value = row.get("world_bank_arrivals")
        row["world_bank_diff"] = None if wb_value is None else row["arrivals"] - wb_value
        row["world_bank_diff_pct"] = None if wb_value in (None, 0) else (row["arrivals"] / wb_value - 1) * 100

    validation = [
        {
            "year": row["year"],
            "mots_total": row["arrivals"],
            "months": row["months"],
            "world_bank_total": row.get("world_bank_arrivals"),
            "difference": row.get("world_bank_diff"),
            "difference_pct": row.get("world_bank_diff_pct"),
            "receipt_total": (receipt_validations or {}).get(row["year"], {}).get("arrivals"),
            "receipt_difference": (
                None
                if row["year"] not in (receipt_validations or {})
                else row["arrivals"] - (receipt_validations or {})[row["year"]]["arrivals"]
            ),
            "receipt_source_file": (receipt_validations or {}).get(row["year"], {}).get("source_file_url"),
            "status": validation_status(row),
        }
        for row in annual
    ]

    return {
        "monthly": monthly,
        "quarterly": quarterly,
        "annual": annual,
        "validation": validation,
    }


def pct_change(value: int | float | None, base: int | float | None) -> float | None:
    if value is None or base is None or base == 0:
        return None
    return (float(value) / float(base) - 1) * 100


def validation_status(row: dict[str, Any]) -> str:
    if not row.get("is_full_year"):
        return "YTD only"
    wb = row.get("world_bank_arrivals")
    if wb is None:
        return "No World Bank annual value"
    diff_pct = abs(row.get("world_bank_diff_pct") or 0)
    if diff_pct <= 0.02:
        return "Match after rounding"
    if diff_pct <= 0.25:
        return "Small rounding/classification gap"
    return "Review"


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if not rows:
        path.write_text("", encoding="utf-8")
        return
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--refresh", action="store_true")
    args = parser.parse_args(argv)

    ensure_dirs()
    source_files, metadata = crawl_mots(refresh=args.refresh)
    monthly_rows, parsed_files = download_and_parse(source_files, refresh=args.refresh)
    trend_rows, trend_source = load_trend_inbound_monthlies(refresh=args.refresh)
    annual_extras, annual_sources = load_annual_extras(refresh=args.refresh)
    receipt_validations, receipt_sources = load_receipt_annual_validations(refresh=args.refresh)
    monthly_rows = select_best_monthly_rows(monthly_rows + trend_rows)
    parsed_files.extend([trend_source, *annual_sources, *receipt_sources])
    wb_values = load_world_bank(refresh=args.refresh)
    derived = build_derived(
        monthly_rows,
        wb_values,
        annual_extras=annual_extras,
        receipt_validations=receipt_validations,
    )

    write_csv(OUT / "tourism_monthly.csv", derived["monthly"])
    write_csv(OUT / "tourism_quarterly.csv", derived["quarterly"])
    write_csv(OUT / "tourism_annual.csv", derived["annual"])
    write_csv(OUT / "validation_annual_worldbank.csv", derived["validation"])
    (OUT / "source_files.json").write_text(
        json.dumps([asdict(source) for source in parsed_files], ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    (OUT / "source_metadata.json").write_text(
        json.dumps(metadata, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    (OUT / "dashboard_data.json").write_text(
        json.dumps({**derived, "metadata": metadata}, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    ok_years = sorted({row["year"] for row in monthly_rows})
    print(f"Parsed monthly rows: {len(monthly_rows)}")
    print(f"Years: {ok_years}")
    print(f"Source files found: {len(source_files)}; parsed ok: {sum(1 for f in parsed_files if f.parse_status == 'ok')}")
    for source in parsed_files:
        if source.parse_status != "ok":
            print(f"WARN parse failed {source.year} {source.file_url}: {source.parse_note}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
